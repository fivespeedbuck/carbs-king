"""Build Carb King's offline exercise catalog from the reviewed name workbook."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


EQUIPMENT = {
    "assisted": "辅助器械", "band": "弹力带", "barbell": "杠铃", "body weight": "自重",
    "bosu ball": "半圆平衡球", "cable": "绳索", "dumbbell": "哑铃", "elliptical machine": "椭圆机",
    "ez barbell": "曲杆杠铃", "hammer": "大锤", "kettlebell": "壶铃", "leverage machine": "悍马机",
    "medicine ball": "药球", "olympic barbell": "奥杆", "resistance band": "弹力带", "roller": "泡沫轴",
    "rope": "战绳", "skierg machine": "滑雪机", "sled machine": "倒蹬机", "smith machine": "史密斯机",
    "stability ball": "健身球", "stationary bike": "动感单车", "stepmill machine": "登阶机", "tire": "轮胎",
    "trap bar": "六角杠铃", "upper body ergometer": "上肢功率车", "weighted": "负重", "wheel roller": "健腹轮",
}

MUSCLES = {
    "abductors": "外展肌群", "abs": "腹肌", "adductors": "内收肌群", "biceps": "肱二头肌",
    "calves": "小腿", "cardiovascular system": "心肺系统", "delts": "三角肌", "forearms": "前臂",
    "glutes": "臀肌", "hamstrings": "腘绳肌", "lats": "背阔肌", "levator scapulae": "肩胛提肌",
    "pectorals": "胸肌", "quads": "股四头肌", "serratus anterior": "前锯肌", "spine": "竖脊肌",
    "traps": "斜方肌", "triceps": "肱三头肌", "upper back": "上背部",
    "abdominals": "腹肌", "ankle stabilizers": "踝稳定肌群", "ankles": "踝部", "back": "背部",
    "brachialis": "肱肌", "chest": "胸肌", "core": "核心肌群", "deltoids": "三角肌",
    "feet": "足部", "grip muscles": "握力肌群", "groin": "腹股沟", "hands": "手部",
    "hip flexors": "髋屈肌", "inner thighs": "大腿内侧", "latissimus dorsi": "背阔肌",
    "lower abs": "下腹", "lower back": "下背部", "obliques": "腹斜肌", "quadriceps": "股四头肌",
    "rear deltoids": "三角肌后束", "rhomboids": "菱形肌", "rotator cuff": "肩袖肌群",
    "shins": "胫骨前肌", "shoulders": "三角肌", "soleus": "比目鱼肌",
    "sternocleidomastoid": "胸锁乳突肌", "trapezius": "斜方肌", "upper chest": "上胸",
    "wrist extensors": "腕伸肌群", "wrist flexors": "腕屈肌群", "wrists": "手腕",
}

GLUTE_DOMINANT_NAME_TERMS = (
    "glute", "hip extension", "hip thrust", "hip lift", "hip internal rotation",
    "bridge", "pull through", "piriformis", "monster walk", "donkey kick",
    "fire hydrant", "clamshell", "frog pump", "kickback", "kettlebell swing",
    "pelvic tilt", "hip abduction", "hip adduction", "lifting (on hip)",
    "reverse hyper",
)

CANONICAL_RECORD_OVERRIDES = json.loads(
    (Path(__file__).resolve().parents[1] / "src" / "exercise_catalog_overrides.json").read_text(encoding="utf-8")
)


def normalize_display_name(value: str) -> str:
    """Remove spreadsheet spacing artefacts and normalize retained abbreviations."""
    name = str(value or "").strip()
    replacements = {
        r"(?i)\bbosu\b": "半圆平衡球",
        r"(?i)\bup\b": "抬起",
        r"(?i)\bto\b": "到",
    }
    for pattern, replacement in replacements.items():
        name = re.sub(pattern, replacement, name)
    name = re.sub(r"(?i)ez", "EZ", name)
    name = re.sub(r"(?i)jm", "JM", name)
    name = re.sub(r"(?i)\bl(?=引体|型)", "L", name)
    name = re.sub(r"(?i)\bt(?=字|杠|杆|划船)", "T", name)
    # Translation workbooks occasionally insert spaces inside Chinese words,
    # e.g. “史密斯 推举”. Exercise names use no word spaces in the UI.
    return re.sub(r"\s+", "", name)


def normalize_equipment_display_name(name: str, equipment: str) -> str:
    """Replace literal machine taxonomy with familiar Chinese gym wording."""
    if equipment == "悍马机" and name.startswith("杠杆式"):
        return f"悍马机{name.removeprefix('杠杆式')}"
    if equipment == "倒蹬机" and name.startswith("雪橇"):
        return name.replace("雪橇", "倒蹬机", 1)
    return name


def category_for(row: dict) -> str:
    category = row["category"]
    target = row["target"]
    name = row["name"].lower()
    if "stretch" in name or "mobility" in name:
        return "拉伸"
    if "warm-up" in name or "warm up" in name:
        return "热身动作"
    if category == "chest": return "胸"
    if category == "back": return "背"
    if category == "shoulders": return "肩"
    if category in {"upper legs", "lower legs"}:
        # The upstream target is the single primary muscle, not an app-level
        # navigation category. A glute target must not move compound squats,
        # leg presses, lunges, deadlifts, or step-ups out of the leg library.
        return "臀部" if target == "glutes" and any(term in name for term in GLUTE_DOMINANT_NAME_TERMS) else "腿"
    if target == "glutes": return "臀部"
    if category == "waist":
        return "核心稳定" if any(word in name for word in ("plank", "dead bug", "bird dog", "pallof")) else "腹部"
    if category == "lower arms": return "小臂"
    if category == "neck": return "颈部"
    if category == "cardio": return "有氧"
    if target == "biceps": return "二头"
    if target == "triceps": return "三头"
    if target == "forearms": return "小臂"
    return "其他"


def subgroup_for(row: dict, category: str) -> str:
    name = row["name"].lower()
    target = row["target"]
    if category == "胸": return "上胸" if "incline" in name else "下胸" if "decline" in name else "中胸"
    if category == "肩": return "后束" if any(word in name for word in ("rear", "reverse", "bent over")) else "中束" if any(word in name for word in ("lateral", "side raise")) else "前束" if "front" in name else "整体"
    if category == "背": return "背阔肌" if target == "lats" else "上背" if target in {"traps", "upper back", "levator scapulae"} else "下背" if target == "spine" else "整体"
    if category == "腿":
        if any(term in name for term in ("squat", "leg press", "lunge", "step-up")):
            return "股四头肌"
        if any(term in name for term in ("deadlift", "good morning", "leg curl")):
            return "腘绳肌"
        return {"quads": "股四头肌", "hamstrings": "腘绳肌", "glutes": "整体", "calves": "小腿", "adductors": "内收肌", "abductors": "外展肌"}.get(target, "整体")
    if category in {"腹部", "核心稳定"}: return "腹斜肌" if "oblique" in name or "side" in name else "下腹" if "leg raise" in name else "核心" if category == "核心稳定" else "上腹"
    if category == "有氧": return EQUIPMENT.get(row["equipment"], "有氧")
    return "整体"


def main(workbook_path: Path, source_path: Path, output_path: Path, override_path: Path | None = None) -> None:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    names = {}
    for serial, source_id, _english, chinese_name in sheet.iter_rows(min_row=2, values_only=True):
        key = str(source_id).zfill(4)
        names[key] = str(chinese_name or "").strip()
    # PowerShell's UTF-8 writer may prepend a BOM; accept both variants so
    # externally maintained translation overrides remain directly usable.
    overrides = json.loads(override_path.read_text(encoding="utf-8-sig")) if override_path else {}
    if not isinstance(overrides, dict):
        raise ValueError("高风险覆盖文件必须是 JSON 对象。")
    names.update({str(key).zfill(4): str(value).strip() for key, value in overrides.items() if str(value).strip()})
    source = json.loads(source_path.read_text(encoding="utf-8"))
    catalog = []
    for row in source:
        source_id = str(row["id"])
        name = normalize_display_name(names.get(source_id, ""))
        if not name:
            continue
        category = category_for(row)
        equipment = EQUIPMENT.get(str(row["equipment"]), str(row["equipment"]))
        name = normalize_equipment_display_name(name, equipment)
        targets = [MUSCLES.get(str(row["target"]), str(row["target"]))]
        targets.extend(MUSCLES.get(str(item), str(item)) for item in row.get("secondary_muscles", []))
        record = {
            "id": f"dataset:{source_id}", "name": name, "category": category,
            "subgroup": subgroup_for(row, category), "equipment": equipment,
            "target_muscles": list(dict.fromkeys(targets)), "cues": list(row.get("instruction_steps", {}).get("zh", [])),
            # Flet resolves asset sources relative to the configured assets directory.
            "mistakes": [], "image": f"exercises/{row['image']}", "gif": f"exercises/gifs/{Path(row['gif_url']).name}",
            "default_weight_kg": None, "default_reps": 10, "default_sets": 4,
            "recording_mode": "cardio" if category == "有氧" else "strength", "distance_enabled": category == "有氧",
            "cardio_metric_fields": [], "aliases": [], "default_duration_seconds": 1200 if category == "有氧" else None,
        }
        record.update(CANONICAL_RECORD_OVERRIDES.get(source_id, {}))
        catalog.append(record)
    output_path.write_text(json.dumps(catalog, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"生成 {len(catalog)} 条动作：{output_path}")


if __name__ == "__main__":
    main(Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3]), Path(sys.argv[4]) if len(sys.argv) > 4 else None)
